from fastapi import FastAPI, UploadFile, File, HTTPException, Query, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
from contextlib import asynccontextmanager
from typing import Optional
import asyncio
import json
import os
from backend import parser
from backend import analytics
from backend import database
from backend import websocket_manager as ws_module
from backend import watcher
from backend import sheets_fetcher

CONFIG_PATH = os.path.join(os.path.dirname(__file__), "config.json")


def _read_config() -> dict:
    try:
        with open(CONFIG_PATH, "r") as f:
            cfg = json.load(f)
    except Exception:
        cfg = {}
    # Environment variables override config.json (used in production on Railway)
    cfg.setdefault("admin_attendance_sheet_id", os.getenv("ADMIN_ATTENDANCE_SHEET_ID", ""))
    cfg.setdefault("admin_leave_sheet_id",      os.getenv("ADMIN_LEAVE_SHEET_ID", ""))
    cfg.setdefault("google_api_key",            os.getenv("GOOGLE_API_KEY", ""))
    cfg.setdefault("admin_attendance_path",     "")
    cfg.setdefault("admin_leave_path",          "")
    # Env vars win over empty config.json values
    if not cfg["admin_attendance_sheet_id"]:
        cfg["admin_attendance_sheet_id"] = os.getenv("ADMIN_ATTENDANCE_SHEET_ID", "")
    if not cfg["admin_leave_sheet_id"]:
        cfg["admin_leave_sheet_id"] = os.getenv("ADMIN_LEAVE_SHEET_ID", "")
    if not cfg["google_api_key"]:
        cfg["google_api_key"] = os.getenv("GOOGLE_API_KEY", "")
    return cfg


def _write_config(data: dict) -> None:
    with open(CONFIG_PATH, "w") as f:
        json.dump(data, f, indent=2)


async def _reload_from_paths(changed_path: str) -> None:
    """Called by file watcher when an Excel file changes."""
    paths = database.get_watch_paths()
    att_path = paths.get("attendance", "")
    leave_path = paths.get("leave", "")

    if not att_path or not os.path.exists(att_path):
        return

    try:
        with open(att_path, "rb") as f:
            att_bytes = f.read()
        leave_bytes = None
        if leave_path and os.path.exists(leave_path):
            with open(leave_path, "rb") as f:
                leave_bytes = f.read()

        att_hash = database.get_file_hash(att_bytes)
        leave_hash = database.get_file_hash(leave_bytes) if leave_bytes else "none"

        # Only re-parse if something actually changed
        cached = database.load_from_cache(att_hash, leave_hash)
        if cached is None:
            dashboard_data = parser.build_dashboard_data(att_bytes, leave_bytes)
            analytics_data = _build_analytics(dashboard_data, att_bytes, leave_bytes)
            database.save_to_cache(att_hash, leave_hash, dashboard_data, analytics_data)

        await ws_module.manager.broadcast({
            "event": "data_updated",
            "changed_file": os.path.basename(changed_path),
        })
    except Exception as e:
        print(f"[watcher] reload error: {e}")


def _is_sheets_mode(cfg: dict) -> bool:
    """True when Google Sheets IDs + API key are configured."""
    return bool(
        cfg.get("admin_attendance_sheet_id", "").strip()
        and cfg.get("google_api_key", "").strip()
    )


async def _reload_from_sheets(att_id: str, leave_id: str, api_key: str) -> bool:
    """
    Fetch latest data from Google Sheets, rebuild cache if content changed.
    Returns True if new data was parsed and cached.
    """
    try:
        dashboard_data = parser.build_dashboard_data_from_google_sheets(
            att_id, leave_id or None, api_key
        )
        if not dashboard_data.get("dates_processed"):
            print(f"[sheets] No dates found. Errors: {dashboard_data.get('errors')}")
            return False

        # Hash the content to detect changes
        content_bytes = json.dumps(dashboard_data, sort_keys=True, default=str).encode()
        content_hash = database.get_file_hash(content_bytes)
        sheets_marker = "gsheets"

        cached = database.load_from_cache(content_hash, sheets_marker)
        if cached is not None:
            return False  # Same content as last fetch

        analytics_data = _build_analytics(dashboard_data, None, None)
        database.save_to_cache(content_hash, sheets_marker, dashboard_data, analytics_data)

        await ws_module.manager.broadcast({"event": "data_updated", "changed_file": "google_sheets"})
        print(f"[sheets] Refreshed — {dashboard_data['employee_count']} employees, "
              f"{len(dashboard_data['dates_processed'])} dates")
        return True
    except Exception as e:
        print(f"[sheets] Reload error: {e}")
        return False


async def _sheets_poll_loop() -> None:
    """Background task: poll Google Sheets every 10 minutes for changes."""
    while True:
        await asyncio.sleep(600)
        cfg = _read_config()
        if _is_sheets_mode(cfg):
            att_id   = sheets_fetcher.extract_sheet_id(cfg["admin_attendance_sheet_id"])
            leave_id = sheets_fetcher.extract_sheet_id(cfg.get("admin_leave_sheet_id", ""))
            api_key  = cfg["google_api_key"].strip()
            await _reload_from_sheets(att_id, leave_id, api_key)


def _build_analytics(dashboard_data: dict, att_bytes: bytes, leave_bytes: bytes | None) -> dict:
    filter_start, filter_end = analytics.get_date_range("all", dashboard_data["dates_processed"])
    filtered = analytics.filter_data_by_date_range(dashboard_data, filter_start, filter_end)
    overall = analytics.get_overall_analytics(filtered, "all")
    leave_a = analytics.get_leave_type_analytics(filtered)
    analytics.enrich_with_week_breakdown(filtered)
    return {
        "period_info": {"period": "all", "start_date": filter_start, "end_date": filter_end,
                        "total_days": len(filtered["dates_processed"])},
        "overall": overall,
        "leave_breakdown": leave_a,
        "filtered_data": filtered,
    }


@asynccontextmanager
async def lifespan(app: FastAPI):
    loop = asyncio.get_event_loop()
    watcher.set_reload_callback(_reload_from_paths, loop)

    cfg = _read_config()

    if _is_sheets_mode(cfg):
        # ── Google Sheets mode ──────────────────────────────────────────────
        # extract_sheet_id handles both full URLs and bare IDs
        att_id   = sheets_fetcher.extract_sheet_id(cfg["admin_attendance_sheet_id"])
        leave_id = sheets_fetcher.extract_sheet_id(cfg.get("admin_leave_sheet_id", ""))
        api_key  = cfg["google_api_key"].strip()
        print(f"[startup] Google Sheets mode — att:{att_id[:12]}... polling every 10 min")
        asyncio.ensure_future(_reload_from_sheets(att_id, leave_id, api_key))
        asyncio.ensure_future(_sheets_poll_loop())
    else:
        # ── Local Excel file mode (legacy / HR mode) ────────────────────────
        att = cfg.get("admin_attendance_path", "").strip()
        leave = cfg.get("admin_leave_path", "").strip()
        if att and os.path.exists(att):
            database.register_watch_path("attendance", att)
            if leave and os.path.exists(leave):
                database.register_watch_path("leave", leave)
            watch_dir = os.path.dirname(os.path.abspath(att))
            watcher.start_watcher(watch_dir)
            asyncio.ensure_future(_reload_from_paths(att))
        else:
            paths = database.get_watch_paths()
            if paths.get("attendance") and os.path.isdir(os.path.dirname(paths["attendance"])):
                watcher.start_watcher(os.path.dirname(paths["attendance"]))

    yield
    watcher.stop_watcher()


app = FastAPI(title="C1X Attendance Transformer", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],   # Vercel frontend domain is set at runtime; wildcard is safe for read-only API
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.post("/api/transform")
async def transform_attendance(
    attendance_file: UploadFile = File(...),
    leave_file: Optional[UploadFile] = File(None),
):
    """Upload raw attendance Excel (+ optional Leave/WFH Excel), get back report."""
    if not attendance_file.filename or not attendance_file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Attendance file must be .xlsx")

    attendance_bytes = await attendance_file.read()
    if len(attendance_bytes) == 0:
        raise HTTPException(status_code=400, detail="Attendance file is empty")

    leave_bytes = None
    if leave_file and leave_file.filename:
        if not leave_file.filename.endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="Leave file must be .xlsx")
        leave_bytes = await leave_file.read()
        if len(leave_bytes) == 0:
            leave_bytes = None

    result = parser.transform_excel(attendance_bytes, leave_bytes)

    if result["output_bytes"] is None:
        raise HTTPException(
            status_code=400,
            detail=result["errors"][0] if result["errors"] else "No data found",
        )

    return Response(
        content=result["output_bytes"],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=Attendance_Report.xlsx",
            "X-Dates-Processed": ",".join(result["dates_processed"]),
            "X-Employee-Count": str(result["employee_count"]),
            "X-Record-Count": str(result["record_count"]),
            "X-Errors": "|".join(result["errors"]) if result["errors"] else "",
            "Access-Control-Expose-Headers": "X-Dates-Processed, X-Employee-Count, X-Record-Count, X-Errors, Content-Disposition",
        },
    )


@app.post("/api/preview")
async def preview_attendance(
    attendance_file: UploadFile = File(...),
    leave_file: Optional[UploadFile] = File(None),
):
    """Upload raw attendance Excel (+ optional Leave/WFH), get back stats only."""
    if not attendance_file.filename or not attendance_file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Attendance file must be .xlsx")

    attendance_bytes = await attendance_file.read()
    if len(attendance_bytes) == 0:
        raise HTTPException(status_code=400, detail="Attendance file is empty")

    leave_bytes = None
    if leave_file and leave_file.filename:
        if not leave_file.filename.endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="Leave file must be .xlsx")
        leave_bytes = await leave_file.read()
        if len(leave_bytes) == 0:
            leave_bytes = None

    result = parser.transform_excel(attendance_bytes, leave_bytes)

    return {
        "dates_processed": result["dates_processed"],
        "employee_count": result["employee_count"],
        "record_count": result["record_count"],
        "errors": result["errors"],
    }


@app.post("/api/dashboard")
async def dashboard_attendance(
    attendance_file: UploadFile = File(...),
    leave_file: Optional[UploadFile] = File(None),
):
    """Upload raw attendance Excel (+ optional Leave/WFH), get back dashboard JSON."""
    if not attendance_file.filename or not attendance_file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Attendance file must be .xlsx")

    attendance_bytes = await attendance_file.read()
    if len(attendance_bytes) == 0:
        raise HTTPException(status_code=400, detail="Attendance file is empty")

    leave_bytes = None
    if leave_file and leave_file.filename:
        if not leave_file.filename.endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="Leave file must be .xlsx")
        leave_bytes = await leave_file.read()
        if len(leave_bytes) == 0:
            leave_bytes = None

    data = parser.build_dashboard_data(attendance_bytes, leave_bytes)

    if not data["dates_processed"]:
        raise HTTPException(
            status_code=400,
            detail=data["errors"][0] if data["errors"] else "No data found",
        )

    return data


@app.get("/")
def root():
    return {"status": "C1X Attendance API is running"}


@app.get("/api/health")
def health():
    return {"status": "ok"}


@app.post("/api/analytics/summary")
async def get_analytics_summary(
    attendance_file: UploadFile = File(...),
    leave_file: Optional[UploadFile] = File(None),
    period: str = Query("today", description="Time period: today, week, month, year, custom"),
    start_date: Optional[str] = Query(None, description="Start date for custom period (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date for custom period (YYYY-MM-DD)"),
):
    """Get comprehensive analytics summary with time-based filtering."""
    if not attendance_file.filename or not attendance_file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Attendance file must be .xlsx")

    attendance_bytes = await attendance_file.read()
    if len(attendance_bytes) == 0:
        raise HTTPException(status_code=400, detail="Attendance file is empty")

    leave_bytes = None
    if leave_file and leave_file.filename:
        if not leave_file.filename.endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="Leave file must be .xlsx")
        leave_bytes = await leave_file.read()
        if len(leave_bytes) == 0:
            leave_bytes = None

    # Build dashboard data
    dashboard_data = parser.build_dashboard_data(attendance_bytes, leave_bytes)
    
    if not dashboard_data["dates_processed"]:
        raise HTTPException(
            status_code=400,
            detail=dashboard_data["errors"][0] if dashboard_data["errors"] else "No data found",
        )

    # Apply date filtering
    filter_start, filter_end = analytics.get_date_range(period, dashboard_data["dates_processed"], start_date, end_date)
    filtered_data = analytics.filter_data_by_date_range(dashboard_data, filter_start, filter_end)
    
    # Get analytics
    overall_analytics = analytics.get_overall_analytics(filtered_data, period)
    leave_analytics = analytics.get_leave_type_analytics(filtered_data)
    analytics.enrich_with_week_breakdown(filtered_data)

    return {
        "period_info": {
            "period": period,
            "start_date": filter_start,
            "end_date": filter_end,
            "total_days": len(filtered_data["dates_processed"])
        },
        "overall": overall_analytics,
        "leave_breakdown": leave_analytics,
        "filtered_data": filtered_data
    }


@app.post("/api/analytics/leaves/{leave_type}")
async def get_employees_by_leave_type(
    leave_type: str,
    attendance_file: UploadFile = File(...),
    leave_file: Optional[UploadFile] = File(None),
    period: str = Query("today", description="Time period: today, week, month, year, custom"),
    start_date: Optional[str] = Query(None, description="Start date for custom period (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date for custom period (YYYY-MM-DD)"),
):
    """Get employees filtered by specific leave type."""
    if not attendance_file.filename or not attendance_file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Attendance file must be .xlsx")

    attendance_bytes = await attendance_file.read()
    leave_bytes = None
    if leave_file and leave_file.filename:
        leave_bytes = await leave_file.read()

    dashboard_data = parser.build_dashboard_data(attendance_bytes, leave_bytes)
    filter_start, filter_end = analytics.get_date_range(period, dashboard_data["dates_processed"], start_date, end_date)
    filtered_data = analytics.filter_data_by_date_range(dashboard_data, filter_start, filter_end)
    
    leave_analytics = analytics.get_leave_type_analytics(filtered_data)
    
    if leave_type not in leave_analytics:
        raise HTTPException(status_code=400, detail=f"Invalid leave type: {leave_type}")
    
    return {
        "leave_type": leave_type,
        "period": period,
        "date_range": {"start": filter_start, "end": filter_end},
        "total_count": leave_analytics[leave_type]["count"],
        "employees": leave_analytics[leave_type]["employees"]
    }


@app.get("/api/admin/config")
async def get_admin_config():
    """Return current admin config (Google Sheets IDs or Excel paths)."""
    cfg = _read_config()
    if _is_sheets_mode(cfg):
        att_id = cfg.get("admin_attendance_sheet_id", "").strip()
        leave_id = cfg.get("admin_leave_sheet_id", "").strip()
        return {
            "mode": "sheets",
            "configured": True,
            "attendance_sheet_id": att_id,
            "leave_sheet_id": leave_id,
            "has_api_key": True,
        }
    att = cfg.get("admin_attendance_path", "").strip()
    leave = cfg.get("admin_leave_path", "").strip()
    return {
        "mode": "excel",
        "configured": bool(att),
        "attendance_path": att,
        "leave_path": leave,
        "attendance_exists": os.path.exists(att) if att else False,
        "leave_exists": os.path.exists(leave) if leave else False,
    }


@app.post("/api/admin/config")
async def save_admin_config(body: dict):
    """
    Save admin config. Accepts either:
      - Google Sheets mode: {attendance_sheet_url, leave_sheet_url (opt), google_api_key}
      - Excel file mode:    {attendance_path, leave_path (opt)}
    """
    att_url = body.get("attendance_sheet_url", "").strip()
    leave_url = body.get("leave_sheet_url", "").strip()
    api_key = body.get("google_api_key", "").strip()
    att_path = body.get("attendance_path", "").strip()
    leave_path = body.get("leave_path", "").strip()

    cfg = _read_config()

    if att_url and api_key:
        # ── Google Sheets mode ──────────────────────────────────────────────
        att_id = sheets_fetcher.extract_sheet_id(att_url)
        leave_id = sheets_fetcher.extract_sheet_id(leave_url) if leave_url else ""

        # Validate connection before saving
        result = sheets_fetcher.validate_connection(att_id, leave_id or None, api_key)
        if not result["ok"]:
            raise HTTPException(status_code=400, detail=result["error"])

        cfg["admin_attendance_sheet_id"] = att_id
        cfg["admin_leave_sheet_id"] = leave_id
        cfg["google_api_key"] = api_key
        _write_config(cfg)

        await _reload_from_sheets(att_id, leave_id, api_key)
        asyncio.ensure_future(_sheets_poll_loop())
        return {"status": "saved", "mode": "sheets", "attendance_sheet_id": att_id}

    elif att_path:
        # ── Excel file mode (legacy) ────────────────────────────────────────
        if not os.path.exists(att_path):
            raise HTTPException(status_code=400, detail=f"File not found: {att_path}")

        cfg["admin_attendance_path"] = att_path
        cfg["admin_leave_path"] = leave_path
        _write_config(cfg)

        database.register_watch_path("attendance", att_path)
        if leave_path and os.path.exists(leave_path):
            database.register_watch_path("leave", leave_path)

        loop = asyncio.get_event_loop()
        watcher.set_reload_callback(_reload_from_paths, loop)
        watcher.start_watcher(os.path.dirname(os.path.abspath(att_path)))

        await _reload_from_paths(att_path)
        return {"status": "saved", "mode": "excel", "watching": True}

    else:
        raise HTTPException(
            status_code=400,
            detail="Provide either (attendance_sheet_url + google_api_key) or attendance_path."
        )


@app.websocket("/ws/live")
async def websocket_live(websocket: WebSocket):
    await ws_module.manager.connect(websocket)
    try:
        while True:
            await websocket.receive_text()  # keep-alive ping/pong
    except WebSocketDisconnect:
        ws_module.manager.disconnect(websocket)


@app.post("/api/set-watch-paths")
async def set_watch_paths(attendance_path: str, leave_path: Optional[str] = None):
    """Register file paths to watch. Backend will auto-reload when these files change."""
    if not os.path.exists(attendance_path):
        raise HTTPException(status_code=400, detail=f"Attendance file not found: {attendance_path}")

    database.register_watch_path("attendance", attendance_path)
    if leave_path:
        database.register_watch_path("leave", leave_path)

    watch_dir = os.path.dirname(os.path.abspath(attendance_path))
    loop = asyncio.get_event_loop()
    watcher.set_reload_callback(_reload_from_paths, loop)
    watcher.start_watcher(watch_dir)

    # Immediately parse and cache
    await _reload_from_paths(attendance_path)

    return {"status": "watching", "watch_dir": watch_dir, "attendance": attendance_path, "leave": leave_path}


@app.post("/api/notify-save")
async def notify_save(body: dict):
    """Called by Excel VBA macro on save. Triggers reload + WebSocket broadcast."""
    file_path = body.get("file_path", "")
    if not file_path:
        raise HTTPException(status_code=400, detail="file_path required")
    await _reload_from_paths(file_path)
    return {"status": "ok", "connections": ws_module.manager.connection_count}


@app.get("/api/cached-dashboard")
async def get_cached_dashboard():
    """Return the last parsed dashboard from SQLite cache (instant, no file upload needed)."""
    cfg = _read_config()

    if _is_sheets_mode(cfg):
        # Always attempt a refresh before serving cached Sheets data so the
        # dashboard reflects current spreadsheet content instead of waiting for
        # the background poll loop.
        att_id   = sheets_fetcher.extract_sheet_id(cfg["admin_attendance_sheet_id"])
        leave_id = sheets_fetcher.extract_sheet_id(cfg.get("admin_leave_sheet_id", ""))
        api_key  = cfg["google_api_key"].strip()
        await _reload_from_sheets(att_id, leave_id, api_key)

        cached = database.get_latest_cache_entry()
        if cached is None:
            raise HTTPException(status_code=500, detail="Failed to load data from Google Sheets.")
        dashboard_data, analytics_data = cached
        return {"dashboard": dashboard_data, "analytics": analytics_data}

    # ── Excel file mode ─────────────────────────────────────────────────────
    paths = database.get_watch_paths()
    att_path = paths.get("attendance", "")
    leave_path = paths.get("leave", "")

    if not att_path or not os.path.exists(att_path):
        raise HTTPException(status_code=404, detail="No data source configured. Use /api/admin/config first.")

    att_hash = database.get_path_hash(att_path)
    leave_hash = database.get_path_hash(leave_path) if leave_path and os.path.exists(leave_path) else "none"

    if att_hash is None:
        raise HTTPException(status_code=404, detail="Attendance file not readable.")

    cached = database.load_from_cache(att_hash, leave_hash)
    if cached is None:
        await _reload_from_paths(att_path)
        att_hash = database.get_path_hash(att_path)
        cached = database.load_from_cache(att_hash, leave_hash or "none")
        if cached is None:
            raise HTTPException(status_code=500, detail="Parsing failed.")

    dashboard_data, analytics_data = cached
    return {"dashboard": dashboard_data, "analytics": analytics_data}


def _rebuild_export_dashboard_data() -> dict:
    """Shared by both export endpoints: rebuilds (or falls back to cached) dashboard data."""
    dashboard_data = None
    cfg = _read_config()

    try:
        if _is_sheets_mode(cfg):
            att_id = sheets_fetcher.extract_sheet_id(cfg["admin_attendance_sheet_id"])
            leave_id = sheets_fetcher.extract_sheet_id(cfg.get("admin_leave_sheet_id", ""))
            dashboard_data = parser.build_dashboard_data_from_google_sheets(
                att_id,
                leave_id or None,
                cfg["google_api_key"],
            )
        else:
            paths = database.get_watch_paths()
            att_path = paths.get("attendance") or cfg.get("admin_attendance_path", "").strip()
            leave_path = paths.get("leave") or cfg.get("admin_leave_path", "").strip()

            if att_path and os.path.exists(att_path):
                with open(att_path, "rb") as f:
                    attendance_bytes = f.read()

                leave_bytes = None
                if leave_path and os.path.exists(leave_path):
                    with open(leave_path, "rb") as f:
                        leave_bytes = f.read()

                dashboard_data = parser.build_dashboard_data(attendance_bytes, leave_bytes)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to rebuild export data: {e}")

    if dashboard_data is None:
        cached = database.get_latest_cache_entry()
        if cached is None:
            raise HTTPException(status_code=404, detail="No dashboard data available. Load the dashboard first.")
        dashboard_data, _ = cached

    if not dashboard_data.get("dates_processed"):
        raise HTTPException(
            status_code=400,
            detail="Export source has no valid attendance dates. Check the attendance workbook tabs and source configuration.",
        )

    return dashboard_data


@app.get("/api/admin/export")
async def export_dashboard_excel():
    """Export the full cached attendance dataset as a month-wise Excel workbook."""
    import io
    from datetime import datetime as _dt
    from collections import OrderedDict
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    dashboard_data = _rebuild_export_dashboard_data()
    employees = dashboard_data.get("employees", [])
    dates = dashboard_data.get("dates_processed", [])

    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    bold_font = Font(bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    fills = {
        "present": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "present_mid": PatternFill(start_color="A7F3D0", end_color="A7F3D0", fill_type="solid"),
        "present_strong": PatternFill(start_color="6EE7B7", end_color="6EE7B7", fill_type="solid"),
        "weekend_worked": PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid"),
        "weekend_worked_strong": PatternFill(start_color="FBBF24", end_color="FBBF24", fill_type="solid"),
        "absent": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "wfh": PatternFill(start_color="CCFBF1", end_color="CCFBF1", fill_type="solid"),
        "leave": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
        "comp_off": PatternFill(start_color="E9D5FF", end_color="E9D5FF", fill_type="solid"),
        "half_leave": PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid"),
        "holiday": PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"),
        "weekend": PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"),
        "week_total": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
        "week_total_extra": PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid"),
        "month_total": PatternFill(start_color="C7F9CC", end_color="C7F9CC", fill_type="solid"),
    }
    worked_font = Font(color="166534", bold=True, size=10)
    weekend_worked_font = Font(color="92400E", bold=True, size=10)
    total_with_extra_font = Font(color="92400E", bold=True, size=10)
    leave_cols = ["WFH", "SL", "CL", "PL", "Comp Off", "Half Day"]

    def status_code(day: dict) -> str:
        st = day.get("status_type", "")
        sub = day.get("leave_subtype", "")
        if st in ("present", "weekend_worked"):
            return "P"
        if st == "absent":
            return "A"
        if st == "wfh":
            return "WFH"
        if st == "half_leave":
            labels = {"half_cl": "1/2CL", "half_sl": "1/2SL", "half_wfh": "1/2WFH", "half_pl": "1/2PL", "half_comp": "1/2CO"}
            return labels.get(sub, "1/2")
        if st == "comp_off":
            return "CO"
        if st == "leave":
            return sub.upper() if sub else "L"
        if st == "holiday":
            return "H"
        if st == "lwd":
            return "LWD"
        if st == "miss_punch":
            return "MP"
        return day.get("status", "")[:4]

    def fill_for(day: dict):
        st = day.get("status_type", "")
        if st == "present":
            mins = day.get("total_minutes") or 0
            if mins >= 600:
                return fills["present_strong"]
            if mins >= 480:
                return fills["present_mid"]
            return fills["present"]
        if st == "weekend_worked":
            mins = day.get("total_minutes") or 0
            return fills["weekend_worked_strong"] if mins >= 480 else fills["weekend_worked"]
        if st == "absent":
            return fills["absent"]
        if st == "wfh":
            return fills["wfh"]
        if st == "leave":
            return fills["leave"]
        if st == "comp_off":
            return fills["comp_off"]
        if st == "half_leave":
            return fills["half_leave"]
        if st in ("holiday", "lwd"):
            return fills["holiday"]
        if day.get("is_weekend"):
            return fills["weekend"]
        return None

    def hhmm(total_minutes: int) -> str:
        hours, mins = divmod(total_minutes, 60)
        return f"{hours:02d}:{mins:02d}"

    def format_total_with_extra(regular_minutes: int, extra_minutes: int) -> str:
        regular = hhmm(regular_minutes) if regular_minutes else "00:00"
        if extra_minutes > 0:
            return f"{regular} + {hhmm(extra_minutes)}"
        total = regular_minutes + extra_minutes
        return hhmm(total) if total else ""

    def style_header_row(ws, row: int, headers: list[str]) -> None:
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

    def build_month_sheet(ws, month_dates: list[str], title: str) -> None:
        month_date_objs = [_dt.strptime(d, "%Y-%m-%d") for d in month_dates]
        week_labels: list[str] = []
        for d_obj in month_date_objs:
            label = f"W{((d_obj.day - 1) // 7) + 1}"
            if label not in week_labels:
                week_labels.append(label)

        # Inline week-total columns appear after each Sunday in the month, and
        # after the final date if the month ends on a non-Sunday (partial trailing week).
        end_of_week_idx_set: set[int] = set()
        for i, d_obj in enumerate(month_date_objs):
            if d_obj.weekday() == 6:
                end_of_week_idx_set.add(i)
        if month_date_objs and (len(month_date_objs) - 1) not in end_of_week_idx_set:
            end_of_week_idx_set.add(len(month_date_objs) - 1)
        end_of_week_indices = sorted(end_of_week_idx_set)

        # Layout: assign each date a column and reserve a column after each end-of-week date
        date_col_map: dict[int, int] = {}
        inline_week_col_map: dict[int, int] = {}
        col_cursor = 5
        for i in range(len(month_dates)):
            date_col_map[i] = col_cursor
            col_cursor += 1
            if i in end_of_week_idx_set:
                inline_week_col_map[i] = col_cursor
                col_cursor += 1

        leave_start_col = col_cursor
        total_summary_start_col = leave_start_col + len(leave_cols)
        week_bucket_start_col = total_summary_start_col + 3
        total_hours_col = week_bucket_start_col + len(week_labels)
        total_cols = total_hours_col

        # Map each date index to the end-of-week index that closes its week
        date_to_segment_idx: dict[int, int] = {}
        seg_pos = 0
        for i in range(len(month_dates)):
            while seg_pos < len(end_of_week_indices) and i > end_of_week_indices[seg_pos]:
                seg_pos += 1
            if seg_pos < len(end_of_week_indices):
                date_to_segment_idx[i] = end_of_week_indices[seg_pos]

        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

        headers = ["S.No", "Emp ID", "Name", "Department"]
        for i, d_obj in enumerate(month_date_objs):
            headers.append(d_obj.strftime("%d\n%a"))
            if i in end_of_week_idx_set:
                headers.append("Week\nTotal")
        headers.extend(leave_cols)
        headers.extend(["Worked Days", "Absent", "Leave Days"])
        headers.extend(f"{label}\nHours" for label in week_labels)
        headers.append("Total Hours")
        style_header_row(ws, 2, headers)

        month_dates_set = set(month_dates)
        employees_with_data = []
        employees_no_data = []
        for emp in employees:
            has_data = any(day.get("date") in month_dates_set for day in emp.get("daily", []))
            (employees_with_data if has_data else employees_no_data).append(emp)

        for idx, emp in enumerate(employees_with_data, 1):
            row = idx + 2
            daily_map = {day["date"]: day for day in emp.get("daily", [])}

            ws.cell(row=row, column=1, value=idx).alignment = center
            ws.cell(row=row, column=2, value=emp["emp_id"]).alignment = center
            ws.cell(row=row, column=3, value=emp["name"])
            ws.cell(row=row, column=4, value=emp["department"])

            wfh_count = 0
            sl_count = 0
            cl_count = 0
            pl_count = 0
            comp_count = 0
            half_count = 0
            present_count = 0
            absent_count = 0
            leave_days_count = 0.0
            weekday_minutes_total = 0
            weekend_minutes_total = 0
            week_minutes = {label: {"weekday": 0, "weekend": 0} for label in week_labels}
            inline_segment_minutes = {eow: {"weekday": 0, "weekend": 0} for eow in end_of_week_indices}

            for index, date_str in enumerate(month_dates):
                col = date_col_map[index]
                day = daily_map.get(date_str)
                cell = ws.cell(row=row, column=col, value=status_code(day) if day else "")
                cell.alignment = center
                cell.border = thin_border

                if not day:
                    continue

                status_type = day.get("status_type", "")
                minutes = day.get("total_minutes") or 0
                is_weekend = bool(day.get("is_weekend"))
                week_label = f"W{((_dt.strptime(date_str, '%Y-%m-%d').day - 1) // 7) + 1}"
                seg_idx = date_to_segment_idx.get(index)

                if status_type in ("present", "weekend_worked") and minutes > 0:
                    cell.value = day.get("total_hhmm") or hhmm(minutes)
                    cell.font = weekend_worked_font if status_type == "weekend_worked" else worked_font

                fill = fill_for(day)
                if fill:
                    cell.fill = fill

                leave_subtype = day.get("leave_subtype", "")
                if status_type in ("present", "weekend_worked"):
                    if is_weekend or status_type == "weekend_worked":
                        weekend_minutes_total += minutes
                        week_minutes[week_label]["weekend"] += minutes
                        if seg_idx is not None:
                            inline_segment_minutes[seg_idx]["weekend"] += minutes
                    else:
                        weekday_minutes_total += minutes
                        week_minutes[week_label]["weekday"] += minutes
                        if seg_idx is not None:
                            inline_segment_minutes[seg_idx]["weekday"] += minutes

                if status_type in ("present", "weekend_worked"):
                    present_count += 1
                elif status_type == "absent":
                    absent_count += 1
                elif status_type == "wfh":
                    wfh_count += 1
                    present_count += 1
                elif status_type == "half_leave":
                    half_count += 0.5
                    leave_days_count += 0.5
                    present_count += 1
                elif status_type == "comp_off":
                    comp_count += 1
                    leave_days_count += 1
                elif status_type == "leave":
                    leave_days_count += 1
                    if leave_subtype == "sl":
                        sl_count += 1
                    elif leave_subtype == "cl":
                        cl_count += 1
                    elif leave_subtype == "pl":
                        pl_count += 1

            # Inline week totals (one column after each Sunday / final partial week)
            for eow_idx, week_col in inline_week_col_map.items():
                weekday_part = inline_segment_minutes[eow_idx]["weekday"]
                weekend_part = inline_segment_minutes[eow_idx]["weekend"]
                cell = ws.cell(
                    row=row,
                    column=week_col,
                    value=format_total_with_extra(weekday_part, weekend_part),
                )
                cell.alignment = center
                cell.border = thin_border
                cell.font = total_with_extra_font if weekend_part else bold_font
                cell.fill = fills["week_total_extra"] if weekend_part else fills["week_total"]

            leave_values = [wfh_count, sl_count, cl_count, pl_count, comp_count, half_count]
            for offset, value in enumerate(leave_values):
                cell = ws.cell(row=row, column=leave_start_col + offset, value=value if value else 0)
                cell.alignment = center
                cell.border = thin_border

            totals = [present_count, absent_count, leave_days_count]
            for offset, value in enumerate(totals):
                cell = ws.cell(row=row, column=total_summary_start_col + offset, value=value)
                cell.alignment = center
                cell.border = thin_border

            for offset, label in enumerate(week_labels):
                weekday_part = week_minutes[label]["weekday"]
                weekend_part = week_minutes[label]["weekend"]
                cell = ws.cell(
                    row=row,
                    column=week_bucket_start_col + offset,
                    value=format_total_with_extra(weekday_part, weekend_part),
                )
                cell.alignment = center
                cell.border = thin_border
                cell.font = total_with_extra_font if weekend_part else bold_font
                cell.fill = fills["week_total_extra"] if weekend_part else fills["week_total"]

            total_cell = ws.cell(
                row=row,
                column=total_hours_col,
                value=format_total_with_extra(weekday_minutes_total, weekend_minutes_total),
            )
            total_cell.alignment = center
            total_cell.border = thin_border
            total_cell.font = total_with_extra_font if weekend_minutes_total else bold_font
            total_cell.fill = fills["month_total"]

            for col in range(1, 5):
                ws.cell(row=row, column=col).border = thin_border

        if employees_no_data:
            no_data_row = len(employees_with_data) + 4
            heading_cell = ws.cell(row=no_data_row, column=1, value="No Data Found")
            heading_cell.font = Font(bold=True, size=11, color="B91C1C")
            ws.merge_cells(start_row=no_data_row, start_column=1, end_row=no_data_row, end_column=total_cols)
            for offset, emp in enumerate(employees_no_data, 1):
                list_row = no_data_row + offset
                ws.cell(row=list_row, column=2, value=emp["emp_id"]).alignment = center
                ws.cell(row=list_row, column=3, value=emp["name"])

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 28
        ws.column_dimensions["D"].width = 20
        for date_col in date_col_map.values():
            ws.column_dimensions[openpyxl.utils.get_column_letter(date_col)].width = 8.5
        for week_col in inline_week_col_map.values():
            ws.column_dimensions[openpyxl.utils.get_column_letter(week_col)].width = 14

        for col in range(leave_start_col, total_summary_start_col):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
        for col in range(total_summary_start_col, week_bucket_start_col):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 13
        for col in range(week_bucket_start_col, total_hours_col + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        ws.freeze_panes = "E3"

    date_objs = [_dt.strptime(d, "%Y-%m-%d") for d in dates]
    dates_by_month: OrderedDict[str, list[str]] = OrderedDict()
    for d_obj, date_str in zip(date_objs, dates):
        dates_by_month.setdefault(d_obj.strftime("%Y-%m"), []).append(date_str)

    if dates_by_month:
        first = True
        for month_dates in dates_by_month.values():
            month_dt = _dt.strptime(month_dates[0], "%Y-%m-%d")
            title = month_dt.strftime("%B %Y Attendance")
            if first:
                ws = wb.active
                ws.title = month_dt.strftime("%b %Y")[:31]
                first = False
            else:
                ws = wb.create_sheet(month_dt.strftime("%b %Y")[:31])
            build_month_sheet(ws, month_dates, title)
    else:
        wb.active.title = "Attendance"

    ws3 = wb.create_sheet("Legend")
    legend = [
        ("05:30", "Worked weekday hours", "D1FAE5"),
        ("08:45", "Worked weekday hours (strong)", "6EE7B7"),
        ("04:30", "Weekend extra hours", "FDE68A"),
        ("A", "Absent", "FEE2E2"),
        ("WFH", "Work From Home", "CCFBF1"),
        ("SL", "Sick Leave", "DBEAFE"),
        ("CL", "Casual Leave", "DBEAFE"),
        ("PL", "Paid Leave", "DBEAFE"),
        ("CO", "Comp Off", "E9D5FF"),
        ("1/2CL", "Half Day CL", "FEF9C3"),
        ("1/2SL", "Half Day SL", "FEF9C3"),
        ("1/2WFH", "Half Day WFH", "FEF9C3"),
        ("1/2PL", "Half Day PL", "FEF9C3"),
        ("1/2CO", "Half Day Comp Off", "FEF9C3"),
        ("08:00 + 04:00", "Weekly or monthly total with weekend extra hours", "FDE68A"),
        ("H", "Holiday", "F3F4F6"),
        ("LWD", "Last Working Day", "F3F4F6"),
        ("MP", "Miss Punch", "FFFFFF"),
    ]
    style_header_row(ws3, 1, ["Code", "Meaning", "Color"])
    for row, (code, meaning, color) in enumerate(legend, 2):
        ws3.cell(row=row, column=1, value=code).border = thin_border
        ws3.cell(row=row, column=2, value=meaning).border = thin_border
        color_cell = ws3.cell(row=row, column=3, value="")
        color_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        color_cell.border = thin_border
    ws3.column_dimensions["A"].width = 10
    ws3.column_dimensions["B"].width = 24
    ws3.column_dimensions["C"].width = 12

    output = io.BytesIO()
    wb.save(output)
    excel_bytes = output.getvalue()

    filename = f"Attendance_Report_{dates[0]}_to_{dates[-1]}.xlsx" if dates else "Attendance_Report.xlsx"
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@app.get("/api/admin/export-employee-summary")
async def export_employee_summary_excel():
    """Export one Excel workbook with a per-employee, month-by-month summary:
    weekly hours worked, month total, and average hours/week. Employees/months
    with no real activity (missing or all-Absent) are excluded and listed under
    a 'Missing Data' section instead."""
    import io
    from datetime import datetime as _dt
    from collections import OrderedDict
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    dashboard_data = _rebuild_export_dashboard_data()
    employees = dashboard_data.get("employees", [])
    dates = dashboard_data.get("dates_processed", [])

    date_objs = [_dt.strptime(d, "%Y-%m-%d") for d in dates]
    dates_by_month: OrderedDict[str, list[str]] = OrderedDict()
    for d_obj, date_str in zip(date_objs, dates):
        dates_by_month.setdefault(d_obj.strftime("%Y-%m"), []).append(date_str)

    def week_label_for(date_str: str) -> str:
        day = _dt.strptime(date_str, "%Y-%m-%d").day
        return f"Week {((day - 1) // 7) + 1}"

    max_weeks = max(
        (len({week_label_for(d) for d in month_dates}) for month_dates in dates_by_month.values()),
        default=0,
    )

    def hhmm(total_minutes: int) -> str:
        hours, mins = divmod(round(total_minutes), 60)
        return f"{hours:02d}:{mins:02d}"

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    name_font = Font(bold=True, size=13)
    bold_font = Font(bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    headers = (
        ["Month", "Employee ID", "Employee Name"]
        + [f"Week {i} Hours" for i in range(1, max_weeks + 1)]
        + ["Total Hours", "Avg Hours/Week"]
    )
    total_cols = len(headers)

    def style_header_row(ws, row: int, values: list[str]) -> None:
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Summary"

    row_cursor = 1
    missing_entries: list[tuple[str, str, str]] = []

    for emp in employees:
        daily_map = {day["date"]: day for day in emp.get("daily", []) if day.get("date")}
        month_rows: list[tuple[str, dict, int]] = []

        for month_key, month_dates in dates_by_month.items():
            month_label = _dt.strptime(month_key, "%Y-%m").strftime("%B %Y")

            has_real_data = False
            week_minutes: "OrderedDict[str, int]" = OrderedDict()
            for date_str in month_dates:
                wl = week_label_for(date_str)
                week_minutes.setdefault(wl, 0)
                day = daily_map.get(date_str)
                if not day:
                    continue
                status_type = day.get("status_type", "")
                if status_type != "absent":
                    has_real_data = True
                if status_type in ("present", "weekend_worked"):
                    week_minutes[wl] += day.get("total_minutes") or 0

            if not has_real_data:
                missing_entries.append((emp["emp_id"], emp["name"], month_label))
                continue

            month_rows.append((month_label, week_minutes, sum(week_minutes.values())))

        if not month_rows:
            continue

        ws.cell(row=row_cursor, column=1, value=emp["name"])
        ws.cell(row=row_cursor, column=1).font = name_font
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=total_cols)
        row_cursor += 1

        style_header_row(ws, row_cursor, headers)
        row_cursor += 1

        for month_label, week_minutes, total_minutes in month_rows:
            ws.cell(row=row_cursor, column=1, value=month_label).border = thin_border
            emp_id_cell = ws.cell(row=row_cursor, column=2, value=emp["emp_id"])
            emp_id_cell.alignment = center
            emp_id_cell.border = thin_border
            ws.cell(row=row_cursor, column=3, value=emp["name"]).border = thin_border

            week_labels_this_month = list(week_minutes.keys())
            for i in range(1, max_weeks + 1):
                col = 3 + i
                cell = ws.cell(row=row_cursor, column=col)
                cell.border = thin_border
                cell.alignment = center
                if i <= len(week_labels_this_month):
                    cell.value = hhmm(week_minutes[week_labels_this_month[i - 1]])

            total_col = 3 + max_weeks + 1
            avg_col = total_col + 1
            total_cell = ws.cell(row=row_cursor, column=total_col, value=hhmm(total_minutes))
            total_cell.alignment = center
            total_cell.border = thin_border
            total_cell.font = bold_font

            num_weeks_this_month = len(week_labels_this_month) or 1
            avg_cell = ws.cell(row=row_cursor, column=avg_col, value=hhmm(total_minutes / num_weeks_this_month))
            avg_cell.alignment = center
            avg_cell.border = thin_border

            row_cursor += 1

        row_cursor += 2

    if missing_entries:
        heading_cell = ws.cell(row=row_cursor, column=1, value="Missing Data")
        heading_cell.font = Font(bold=True, size=13, color="B91C1C")
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=total_cols)
        row_cursor += 1

        style_header_row(ws, row_cursor, ["Employee ID", "Employee Name", "Month"])
        row_cursor += 1

        for emp_id, name, month_label in missing_entries:
            id_cell = ws.cell(row=row_cursor, column=1, value=emp_id)
            id_cell.alignment = center
            id_cell.border = thin_border
            ws.cell(row=row_cursor, column=2, value=name).border = thin_border
            ws.cell(row=row_cursor, column=3, value=month_label).border = thin_border
            row_cursor += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 26
    for i in range(1, max_weeks + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(3 + i)].width = 13
    ws.column_dimensions[openpyxl.utils.get_column_letter(3 + max_weeks + 1)].width = 14
    ws.column_dimensions[openpyxl.utils.get_column_letter(3 + max_weeks + 2)].width = 16

    output = io.BytesIO()
    wb.save(output)
    excel_bytes = output.getvalue()

    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="Employee_Summary_Report.xlsx"',
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@app.get("/api/admin/underperformers")
async def get_underperformers_from_cache(
    period: str = Query("month"),
    threshold_hours: int = Query(40),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
):
    """Get underperforming employees from cached dashboard data (no file upload needed)."""
    cached = database.get_latest_cache_entry()
    if cached is None:
        raise HTTPException(status_code=404, detail="No cached data. Load the dashboard first.")
    dashboard_data, _ = cached
    if start_date and end_date:
        filtered = analytics.filter_data_by_date_range(dashboard_data, start_date, end_date)
    else:
        fs, fe = analytics.get_date_range(period, dashboard_data["dates_processed"])
        filtered = analytics.filter_data_by_date_range(dashboard_data, fs, fe)
    return analytics.get_underperforming_employees(filtered, period, threshold_hours)


@app.post("/api/sheets/refresh")
async def refresh_sheets():
    """Manually trigger a re-fetch from Google Sheets (force bypass cache)."""
    cfg = _read_config()
    if not _is_sheets_mode(cfg):
        raise HTTPException(status_code=400, detail="Not in Google Sheets mode.")
    att_id   = sheets_fetcher.extract_sheet_id(cfg["admin_attendance_sheet_id"])
    leave_id = sheets_fetcher.extract_sheet_id(cfg.get("admin_leave_sheet_id", ""))
    api_key  = cfg["google_api_key"].strip()
    updated = await _reload_from_sheets(att_id, leave_id, api_key)
    return {"refreshed": updated}


@app.get("/api/watch-status")
async def watch_status():
    """Returns current watch path config and last-parsed info."""
    paths = database.get_watch_paths()
    info = database.get_last_parsed_info()
    return {
        "watching": bool(paths.get("attendance")),
        "paths": paths,
        "last_parsed": info,
        "connected_dashboards": ws_module.manager.connection_count,
    }


@app.post("/api/analytics/underperformers")
async def get_underperforming_employees_endpoint(
    attendance_file: UploadFile = File(...),
    leave_file: Optional[UploadFile] = File(None),
    period: str = Query("week", description="Time period: week, month, year"),
    threshold_hours: int = Query(40, description="Minimum hours threshold"),
    start_date: Optional[str] = Query(None, description="Start date for custom period (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date for custom period (YYYY-MM-DD)"),
):
    """Get employees working below threshold hours."""
    if not attendance_file.filename or not attendance_file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Attendance file must be .xlsx")

    attendance_bytes = await attendance_file.read()
    leave_bytes = None
    if leave_file and leave_file.filename:
        leave_bytes = await leave_file.read()

    dashboard_data = parser.build_dashboard_data(attendance_bytes, leave_bytes)
    
    # Apply custom date filtering if provided
    if start_date and end_date:
        filtered_data = analytics.filter_data_by_date_range(dashboard_data, start_date, end_date)
    else:
        filter_start, filter_end = analytics.get_date_range(period, dashboard_data["dates_processed"])
        filtered_data = analytics.filter_data_by_date_range(dashboard_data, filter_start, filter_end)
    
    underperformers = analytics.get_underperforming_employees(filtered_data, period, threshold_hours)
    
    return underperformers
