"""
Transforms raw attendance Excel + Leave/WFH Excel into a clean combined report.

Attendance file (per sheet named DD-MM-YYYY):
  Row 8+: No. | Employee ID | First Name | Department | Weekday | First Punch | Last Punch | Total Time | IN Temp | OUT Temp

Leave/WFH file (per sheet named "Jan", "Feb", etc.):
  Row 2+: S.No | Emp ID | Name | Day1 | Day2 | ... | Day31 | WFH Count | SL | CL | PL
  Values: WFH, CL, SL, PL, COMP OFF, 1/2CL, 1/2SL, Holiday, LWD, or empty (present)

Combined status logic:
  1. Leave file says WFH/CL/SL/PL/COMP OFF/Holiday → use that status
  2. Attendance has punch → PRESENT with In/Out/Total
  3. Weekend + no data → HOLIDAY (gray)
  4. Weekday + no data → ABSENT (red)
  5. Weekend + punch → Weekend worker (yellow highlight)

Output report:
  S.No | EMP ID | Employee Name | Department
  | [Date columns: Status | In | Out | Total] ...
  | Weekday Hrs | Weekend Hrs | Total Hrs | Working Days | Weekend Days | WFH Days | Leave Days
  | Week 1 Total | Week 2 Total | ... (weekly totals shown at the end)
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import OrderedDict
from datetime import datetime, time, timedelta
from typing import Optional
from calendar import monthrange
import io
import re


# ── Month name mapping ──
MONTH_MAP = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}

# Leave status codes
LEAVE_STATUSES = {"WFH", "CL", "SL", "PL", "COMP OFF", "1/2CL", "1/2SL", "1/2WFH", "1/2PL", "1/2COMP", "HOLIDAY", "LWD"}

# Half-day variants: "1/2 CL", "HALF-CL", "½ CL", "0.5 WFH", "1/2COMP OFF", etc.
_HALF_PATTERN = re.compile(
    r"^(?:1\s*/\s*2|HALF|0?\.5|½)\s*[-._/ ]*\s*(CL|SL|PL|WFH|COMP\s*OFF|COMPOFF|COMP)$"
)
_COMP_OFF_PATTERN = re.compile(r"^COMP[-_\s]*OFF$")


def _normalize_leave_status(raw) -> str:
    """
    Canonicalize a raw leave-cell value to one of LEAVE_STATUSES.
    Returns "" for blank/unknown cells (which are treated as "no leave marker").
    Handles case, whitespace, hyphens, underscores, and the Unicode ½ symbol.
    """
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s:
        return ""
    s = s.upper().replace("½", "1/2")
    if s in ("NAN", "NONE", "N/A", "-", "--"):
        return ""

    # Half-day variants — must check before plain codes
    m = _HALF_PATTERN.match(s)
    if m:
        suffix = re.sub(r"\s+", "", m.group(1))
        if suffix in ("COMPOFF", "COMP"):
            return "1/2COMP"
        return f"1/2{suffix}"

    # Comp Off variants
    if _COMP_OFF_PATTERN.match(s) or s == "COMP":
        return "COMP OFF"

    # Plain codes
    if s in {"WFH", "CL", "SL", "PL", "HOLIDAY", "LWD"}:
        return s

    return ""


def parse_time_value(value) -> Optional[str]:
    """Convert Excel time cell to HH:MM string."""
    if value is None:
        return None
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, timedelta):
        total_seconds = int(value.total_seconds())
        if total_seconds < 0:
            return None
        hours, remainder = divmod(total_seconds, 3600)
        minutes = remainder // 60
        return f"{hours:02d}:{minutes:02d}"
    s = str(value).strip()
    if re.match(r"^\d{1,2}:\d{2}$", s):
        parts = s.split(":")
        return f"{int(parts[0]):02d}:{int(parts[1]):02d}"
    # HH:MM:SS — returned by Google Sheets FORMATTED_VALUE for time columns
    if re.match(r"^\d{1,2}:\d{2}:\d{2}$", s):
        parts = s.split(":")
        return f"{int(parts[0]):02d}:{int(parts[1]):02d}"
    return None


def time_to_minutes(time_str: Optional[str]) -> Optional[int]:
    """Convert HH:MM string to total minutes."""
    if not time_str:
        return None
    try:
        parts = time_str.split(":")
        return int(parts[0]) * 60 + int(parts[1])
    except (ValueError, IndexError):
        return None


def minutes_to_hhmm(minutes: Optional[int]) -> str:
    """Convert total minutes to HH:MM string."""
    if minutes is None or minutes <= 0:
        return ""
    h, m = divmod(minutes, 60)
    return f"{h:02d}:{m:02d}"


def is_weekend(date_str: str) -> bool:
    """Check if a YYYY-MM-DD date is Saturday(5) or Sunday(6)."""
    return datetime.strptime(date_str, "%Y-%m-%d").weekday() >= 5


def parse_leave_file(file_bytes: bytes, year: int) -> tuple[dict, dict, list]:
    """
    Parse Leave/WFH Excel file.

    Args:
        file_bytes: raw bytes of the Leave/WFH .xlsx
        year: the year to assign dates (derived from attendance file dates)

    Returns:
        (leave_records, leave_employees, errors)
        leave_records: {emp_id: {date_str: status}} e.g. {"191": {"2026-02-01": "WFH"}}
        leave_employees: {emp_id: {"name": ..., "department": ...}}
        errors: list of warning strings
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    errors = []
    leave_records: dict[str, dict[str, str]] = {}
    leave_employees: dict[str, dict] = {}

    for sheet_name in wb.sheetnames:
        # Match month name
        clean_name = sheet_name.strip().lower()
        month_num = MONTH_MAP.get(clean_name)
        if month_num is None:
            errors.append(f"Leave file: skipped sheet '{sheet_name}' (not a recognized month)")
            continue

        ws = wb[sheet_name]
        days_in_month = monthrange(year, month_num)[1]

        # Find the header row to locate day columns
        # Expected: S.No | Emp ID | Name | 1 | 2 | 3 | ... | 31 | WFH | SL | CL | PL
        # We look for the row where columns contain day numbers 1, 2, 3...
        header_row_idx = None
        day_col_start = None

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
            if not row:
                continue
            for col_idx, val in enumerate(row):
                if val is not None:
                    try:
                        if int(float(str(val))) == 1:
                            # Check if next cell is 2
                            if col_idx + 1 < len(row) and row[col_idx + 1] is not None:
                                try:
                                    if int(float(str(row[col_idx + 1]))) == 2:
                                        header_row_idx = row_idx
                                        day_col_start = col_idx
                                        break
                                except (ValueError, TypeError):
                                    pass
                    except (ValueError, TypeError):
                        pass
            if header_row_idx is not None:
                break

        if header_row_idx is None:
            # Fallback: assume header is row 1, day columns start at index 3
            header_row_idx = 1
            day_col_start = 3

        # Parse data rows (start after header)
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not row or len(row) < day_col_start + 1:
                continue

            # Emp ID is column B (index 1)
            emp_id_raw = row[1]
            if emp_id_raw is None:
                continue
            emp_id_str = str(emp_id_raw).strip()
            if not emp_id_str or emp_id_str.lower() in ("total", "grand total", "nan"):
                continue
            try:
                int(float(emp_id_str))
            except (ValueError, TypeError):
                continue

            emp_id = str(int(float(emp_id_str)))
            emp_name = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            if not emp_name:
                continue  # skip rows with no employee name

            if emp_id not in leave_employees:
                leave_employees[emp_id] = {"name": emp_name, "department": ""}

            if emp_id not in leave_records:
                leave_records[emp_id] = {}

            # Read day columns
            for day in range(1, days_in_month + 1):
                col_idx = day_col_start + (day - 1)
                if col_idx >= len(row):
                    break

                cell_val = row[col_idx]
                status = _normalize_leave_status(cell_val)
                if status in LEAVE_STATUSES:
                    date_str = f"{year}-{month_num:02d}-{day:02d}"
                    leave_records[emp_id][date_str] = status

    wb.close()
    return leave_records, leave_employees, errors


def transform_excel(attendance_bytes: bytes, leave_bytes: Optional[bytes] = None) -> dict:
    """
    Parse raw attendance Excel (and optionally Leave/WFH Excel) and produce a clean report.

    Returns dict with:
      - 'output_bytes': bytes of the generated .xlsx report
      - 'dates_processed': list of date strings
      - 'employee_count': number of unique employees
      - 'record_count': total attendance records
      - 'errors': list of warning/error messages
    """
    wb = openpyxl.load_workbook(io.BytesIO(attendance_bytes), data_only=True)
    errors = []

    # ── Step 1: Parse attendance sheets ──
    employees: dict[str, dict] = {}
    records: dict[str, dict[str, dict]] = {}
    all_dates: list[str] = []

    for sheet_name in wb.sheetnames:
        try:
            sheet_date = datetime.strptime(sheet_name.strip(), "%d-%m-%Y")
            date_str = sheet_date.strftime("%Y-%m-%d")
        except ValueError:
            errors.append(f"Skipped sheet '{sheet_name}': not a valid DD-MM-YYYY date")
            continue

        all_dates.append(date_str)
        ws = wb[sheet_name]

        for row in ws.iter_rows(min_row=8, values_only=True):
            if not row or len(row) < 8:
                continue

            emp_id_raw = row[1]
            if emp_id_raw is None:
                continue
            emp_id_str = str(emp_id_raw).strip()
            if not emp_id_str or emp_id_str.lower() in ("total", "grand total"):
                continue
            try:
                int(float(emp_id_str))
            except (ValueError, TypeError):
                continue

            emp_id = str(int(float(emp_id_str)))
            emp_name = str(row[2]).strip() if row[2] else ""
            if not emp_name:
                continue  # skip rows with no employee name
            department = str(row[3]).strip() if row[3] else "Unknown"

            in_time = parse_time_value(row[5])
            out_time = parse_time_value(row[6])
            total_time_str = parse_time_value(row[7])

            total_minutes = time_to_minutes(total_time_str)
            if total_minutes is None and in_time and out_time:
                in_min = time_to_minutes(in_time)
                out_min = time_to_minutes(out_time)
                if in_min is not None and out_min is not None and out_min > in_min:
                    total_minutes = out_min - in_min

            if emp_id not in employees:
                employees[emp_id] = {"name": emp_name, "department": department}

            if emp_id not in records:
                records[emp_id] = {}
            records[emp_id][date_str] = {
                "in_time": in_time or "",
                "out_time": out_time or "",
                "total_minutes": total_minutes,
            }

    wb.close()
    all_dates = sorted(set(all_dates))

    if not all_dates:
        return {
            "output_bytes": None,
            "dates_processed": [],
            "employee_count": 0,
            "record_count": 0,
            "errors": errors or ["No valid date sheets found in the file"],
        }

    # ── Step 2: Parse leave file (if provided) ──
    leave_records: dict[str, dict[str, str]] = {}
    leave_employees: dict[str, dict] = {}

    if leave_bytes:
        # Derive year from attendance dates
        year = datetime.strptime(all_dates[0], "%Y-%m-%d").year
        leave_records, leave_employees, leave_errors = parse_leave_file(leave_bytes, year)
        errors.extend(leave_errors)

        # Merge employee info: prefer leave file names (cleaner), keep department from attendance
        for emp_id, leave_emp in leave_employees.items():
            if emp_id in employees:
                # Keep attendance department, use leave name if cleaner
                if leave_emp["name"] and leave_emp["name"] != "Unknown":
                    employees[emp_id]["name"] = leave_emp["name"]
            else:
                # Employee only in leave file (e.g. full WFH, never punched in)
                employees[emp_id] = {"name": leave_emp["name"], "department": leave_emp.get("department", "Unknown")}

    # Pre-compute which dates are weekends
    weekend_flags = {d: is_weekend(d) for d in all_dates}

    # ── Step 3: Group dates into weeks ──
    weeks = OrderedDict()  # {week_label: [date_str, ...]}
    date_to_week = {}       # {date_str: week_label}
    seen_iso_weeks = {}     # {(iso_year, iso_week): week_label}

    for date_str in all_dates:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        iso_year, iso_week, _ = dt.isocalendar()
        iso_key = (iso_year, iso_week)
        if iso_key not in seen_iso_weeks:
            week_label = f"Week {len(weeks) + 1}"
            seen_iso_weeks[iso_key] = week_label
            weeks[week_label] = []
        week_label = seen_iso_weeks[iso_key]
        weeks[week_label].append(date_str)
        date_to_week[date_str] = week_label

    num_weeks = len(weeks)

    # ── Build column layout map ──
    # For each date: 3 columns (In, Out, Total)
    # Summary columns next
    # Weekly total columns at the end (not between weeks)
    col_offset = 5  # first date column starts at col 5 (after S.No, EMP ID, Name, Dept)
    date_col_map = {}       # {date_str: start_col} (the In Time column for that date)
    week_total_col_map = OrderedDict()  # {week_label: col}

    current_col = col_offset
    for date_str in all_dates:
        date_col_map[date_str] = current_col
        current_col += 3  # In, Out, Total

    # ── Summary column layout ──
    num_summary_cols = 7
    summary_start = current_col
    current_col += num_summary_cols

    # ── Weekly totals at the end ──
    for week_label in weeks.keys():
        week_total_col_map[week_label] = current_col
        current_col += 1  # weekly total column

    # ── Step 4: Build output Excel ──
    out_wb = openpyxl.Workbook()
    ws = out_wb.active
    ws.title = "Attendance Report"

    # ── Styles ──
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    weekday_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    weekend_header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    weekend_cell_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    weekend_worked_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    miss_punch_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # yellow
    miss_punch_font = Font(color="CC0000", bold=True)
    holiday_font = Font(color="808080", italic=True)
    absent_font = Font(color="CC0000", italic=True)
    # Leave status fills
    wfh_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # light green
    wfh_font = Font(color="375623", bold=True)
    leave_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")  # light pink
    leave_font = Font(color="C62828", italic=True)
    holiday_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")  # light indigo
    comp_off_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")  # light orange
    comp_off_font = Font(color="E65100", italic=True)
    half_day_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # light yellow
    half_day_font = Font(color="F57F17")
    # Weekly total column styles
    week_total_header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    week_total_cell_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    # Summary columns
    total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    weekend_total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    grand_total_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    wfh_total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    leave_total_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Summary + weekly totals layout ──
    total_report_cols = current_col - 1

    # ── Row 1: Title ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_report_cols)
    title_cell = ws.cell(row=1, column=1, value="C1X ATTENDANCE REPORT")
    title_cell.font = Font(bold=True, size=14, color="2F5496")
    title_cell.alignment = Alignment(horizontal="center")

    # ── Row 2: Date range ──
    if all_dates:
        start_d = datetime.strptime(all_dates[0], "%Y-%m-%d").strftime("%d %b %Y")
        end_d = datetime.strptime(all_dates[-1], "%Y-%m-%d").strftime("%d %b %Y")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_report_cols)
        ws.cell(row=2, column=1, value=f"Period: {start_d} to {end_d}").font = Font(size=10, italic=True)

    # ── Row 3: Legend ──
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_report_cols)
    legend_parts = [
        "Gray = Weekend/Holiday",
        "Yellow = Weekend worker / Miss punch",
        "Green = WFH",
        "Pink = CL/SL/PL",
        "Orange = Comp Off",
        "Red = Absent",
    ]
    legend_cell = ws.cell(row=3, column=1, value="Legend:  " + "  |  ".join(legend_parts))
    legend_cell.font = Font(size=9, italic=True, color="555555")

    # ── Row 4-5: Headers ──
    header_row = 4
    sub_header_row = 5

    # Fixed columns
    fixed_headers = ["S.No", "EMP ID", "Employee Name", "Department"]
    for col_idx, header_text in enumerate(fixed_headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.merge_cells(start_row=header_row, start_column=col_idx, end_row=sub_header_row, end_column=col_idx)

    # Date columns + weekly total columns
    for date_str in all_dates:
        start_col = date_col_map[date_str]
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        day_name = dt.strftime("%a")
        display_date = dt.strftime("%d-%b-%Y") + f" ({day_name})"
        is_wkend = weekend_flags[date_str]

        fill = weekend_header_fill if is_wkend else weekday_header_fill

        ws.merge_cells(start_row=header_row, start_column=start_col, end_row=header_row, end_column=start_col + 2)
        date_cell = ws.cell(row=header_row, column=start_col, value=display_date)
        date_cell.font = header_font
        date_cell.fill = fill
        date_cell.alignment = center_align
        date_cell.border = thin_border

        for sub_idx, sub_name in enumerate(["In Time", "Out Time", "Total Hrs"]):
            cell = ws.cell(row=sub_header_row, column=start_col + sub_idx, value=sub_name)
            cell.font = Font(bold=True, size=9, color="FFFFFF")
            cell.fill = fill
            cell.alignment = center_align
            cell.border = thin_border

    # ── Summary columns ──
    summary_headers = [
        ("Weekday\nHours", total_fill),
        ("Weekend\nHours", weekend_total_fill),
        ("Total\nHours", grand_total_fill),
        ("Working\nDays", total_fill),
        ("Weekend\nDays", weekend_total_fill),
        ("WFH\nDays", wfh_total_fill),
        ("Leave\nDays", leave_total_fill),
    ]
    for i, (label, fill) in enumerate(summary_headers):
        col = summary_start + i
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = wrap_center
        cell.border = thin_border
        ws.merge_cells(start_row=header_row, start_column=col, end_row=sub_header_row, end_column=col)

    # Weekly total column headers (at the end of the table)
    for week_label, wt_col in week_total_col_map.items():
        cell = ws.cell(row=header_row, column=wt_col, value=f"{week_label}\nTotal")
        cell.font = header_font
        cell.fill = week_total_header_fill
        cell.alignment = wrap_center
        cell.border = thin_border
        ws.merge_cells(start_row=header_row, start_column=wt_col, end_row=sub_header_row, end_column=wt_col)

    # ── Data Rows ──
    sorted_emps = sorted(employees.keys(), key=lambda eid: int(eid))
    data_start_row = sub_header_row + 1
    record_count = 0

    for row_idx, emp_id in enumerate(sorted_emps):
        row_num = data_start_row + row_idx
        emp = employees[emp_id]

        ws.cell(row=row_num, column=1, value=row_idx + 1).alignment = center_align
        ws.cell(row=row_num, column=2, value=emp_id).alignment = center_align
        ws.cell(row=row_num, column=3, value=emp["name"])
        ws.cell(row=row_num, column=4, value=emp["department"])

        weekday_minutes = 0
        weekend_minutes = 0
        weekday_days = 0
        weekend_days = 0
        wfh_days = 0
        leave_days = 0
        weekly_minutes = {week_label: 0 for week_label in weeks.keys()}

        for date_idx, date_str in enumerate(all_dates):
            start_col = date_col_map[date_str]
            is_wkend = weekend_flags[date_str]
            rec = records.get(emp_id, {}).get(date_str)
            leave_status = leave_records.get(emp_id, {}).get(date_str)

            has_punch = rec and (rec["in_time"] or rec["out_time"])
            day_minutes = 0

            # ── Determine what to show for this cell ──

            if leave_status == "WFH":
                # WFH — show status, count as working day
                record_count += 1
                wfh_days += 1
                day_minutes = 480  # 8 hours default for WFH
                weekday_minutes += day_minutes
                cell = ws.cell(row=row_num, column=start_col, value="WFH")
                cell.font = wfh_font
                ws.merge_cells(start_row=row_num, start_column=start_col, end_row=row_num, end_column=start_col + 2)
                for c in range(3):
                    ws.cell(row=row_num, column=start_col + c).fill = wfh_fill

            elif leave_status in ("CL", "SL", "PL"):
                # Full day leave
                leave_days += 1
                cell = ws.cell(row=row_num, column=start_col, value=leave_status)
                cell.font = leave_font
                ws.merge_cells(start_row=row_num, start_column=start_col, end_row=row_num, end_column=start_col + 2)
                for c in range(3):
                    ws.cell(row=row_num, column=start_col + c).fill = leave_fill

            elif leave_status == "COMP OFF":
                leave_days += 1
                cell = ws.cell(row=row_num, column=start_col, value="Comp Off")
                cell.font = comp_off_font
                ws.merge_cells(start_row=row_num, start_column=start_col, end_row=row_num, end_column=start_col + 2)
                for c in range(3):
                    ws.cell(row=row_num, column=start_col + c).fill = comp_off_fill

            elif leave_status in ("1/2CL", "1/2SL", "1/2WFH", "1/2PL", "1/2COMP"):
                # Half day leave — show status + punch data if exists
                leave_days += 0.5
                half_labels = {"1/2CL": "½CL", "1/2SL": "½SL", "1/2WFH": "½WFH", "1/2PL": "½PL", "1/2COMP": "½COMP"}
                half_label = half_labels.get(leave_status, leave_status)
                if has_punch:
                    record_count += 1
                    mins = rec["total_minutes"] if rec["total_minutes"] and rec["total_minutes"] > 0 else 0
                    day_minutes = mins
                    weekday_minutes += mins
                    weekday_days += 1
                    ws.cell(row=row_num, column=start_col, value=rec["in_time"] or "")
                    ws.cell(row=row_num, column=start_col + 1, value=rec["out_time"] or "")
                    ws.cell(row=row_num, column=start_col + 2, value=f"{minutes_to_hhmm(mins)} ({half_label})")
                else:
                    cell = ws.cell(row=row_num, column=start_col, value=half_label)
                    cell.font = half_day_font
                    ws.merge_cells(start_row=row_num, start_column=start_col, end_row=row_num, end_column=start_col + 2)
                for c in range(3):
                    ws.cell(row=row_num, column=start_col + c).fill = half_day_fill

            elif leave_status == "HOLIDAY":
                # Company holiday
                cell = ws.cell(row=row_num, column=start_col, value="Holiday")
                cell.font = holiday_font
                ws.merge_cells(start_row=row_num, start_column=start_col, end_row=row_num, end_column=start_col + 2)
                for c in range(3):
                    ws.cell(row=row_num, column=start_col + c).fill = holiday_fill

            elif leave_status == "LWD":
                # Last Working Day — treat like present if has punch
                if has_punch:
                    record_count += 1
                    mins = rec["total_minutes"] if rec["total_minutes"] and rec["total_minutes"] > 0 else 0
                    day_minutes = mins
                    weekday_minutes += mins
                    weekday_days += 1
                    ws.cell(row=row_num, column=start_col, value=rec["in_time"] or "")
                    ws.cell(row=row_num, column=start_col + 1, value=rec["out_time"] or "")
                    ws.cell(row=row_num, column=start_col + 2, value=minutes_to_hhmm(mins))
                else:
                    cell = ws.cell(row=row_num, column=start_col, value="LWD")
                    cell.font = leave_font
                    ws.merge_cells(start_row=row_num, start_column=start_col, end_row=row_num, end_column=start_col + 2)

            elif has_punch:
                # No leave status, has punch data → PRESENT
                record_count += 1
                ws.cell(row=row_num, column=start_col, value=rec["in_time"] or "")
                ws.cell(row=row_num, column=start_col + 1, value=rec["out_time"] or "")
                ws.cell(row=row_num, column=start_col + 2, value=minutes_to_hhmm(rec["total_minutes"]))

                mins = rec["total_minutes"] if rec["total_minutes"] and rec["total_minutes"] > 0 else 0
                day_minutes = mins

                if is_wkend:
                    # Weekend worker → yellow
                    weekend_minutes += mins
                    weekend_days += 1
                    for c in range(3):
                        ws.cell(row=row_num, column=start_col + c).fill = weekend_worked_fill
                else:
                    weekday_minutes += mins
                    weekday_days += 1

            elif is_wkend:
                # Weekend, no punch, no leave → Holiday
                cell = ws.cell(row=row_num, column=start_col, value="Holiday")
                cell.font = holiday_font
                ws.merge_cells(start_row=row_num, start_column=start_col, end_row=row_num, end_column=start_col + 2)
                for c in range(3):
                    ws.cell(row=row_num, column=start_col + c).fill = weekend_cell_fill

            else:
                # Weekday, no punch, no leave → Absent
                cell = ws.cell(row=row_num, column=start_col, value="Absent")
                cell.font = absent_font
                ws.merge_cells(start_row=row_num, start_column=start_col, end_row=row_num, end_column=start_col + 2)

            # Apply alignment + border
            for c in range(3):
                cell = ws.cell(row=row_num, column=start_col + c)
                cell.alignment = center_align
                cell.border = thin_border

            if day_minutes and day_minutes > 0:
                week_label = date_to_week[date_str]
                weekly_minutes[week_label] += day_minutes

        # ── Summary cells ──
        grand_total = weekday_minutes + weekend_minutes

        summary_data = [
            (minutes_to_hhmm(weekday_minutes), total_fill),
            (minutes_to_hhmm(weekend_minutes), weekend_total_fill),
            (minutes_to_hhmm(grand_total), grand_total_fill),
            (weekday_days, total_fill),
            (weekend_days, weekend_total_fill),
            (wfh_days, wfh_total_fill),
            (leave_days if leave_days == int(leave_days) else leave_days, leave_total_fill),
        ]

        for i, (value, fill) in enumerate(summary_data):
            c = ws.cell(row=row_num, column=summary_start + i, value=value)
            c.font = Font(bold=True)
            c.fill = fill
            c.alignment = center_align
            c.border = thin_border

        # Weekly totals at the end
        for week_label, wt_col in week_total_col_map.items():
            value = minutes_to_hhmm(weekly_minutes.get(week_label, 0))
            c = ws.cell(row=row_num, column=wt_col, value=value)
            c.font = Font(bold=True)
            c.fill = week_total_cell_fill
            c.alignment = center_align
            c.border = thin_border

        # Border for fixed columns
        for c_idx in range(1, 5):
            ws.cell(row=row_num, column=c_idx).border = thin_border

    # ── Column widths ──
    ws.column_dimensions[get_column_letter(1)].width = 6     # S.No
    ws.column_dimensions[get_column_letter(2)].width = 10    # EMP ID
    ws.column_dimensions[get_column_letter(3)].width = 25    # Employee Name
    ws.column_dimensions[get_column_letter(4)].width = 15    # Department

    for date_str in all_dates:
        start_col = date_col_map[date_str]
        for c in range(3):
            ws.column_dimensions[get_column_letter(start_col + c)].width = 10

    for i in range(num_summary_cols):
        ws.column_dimensions[get_column_letter(summary_start + i)].width = 13

    for wt_col in week_total_col_map.values():
        ws.column_dimensions[get_column_letter(wt_col)].width = 12

    # Freeze panes
    ws.freeze_panes = ws.cell(row=data_start_row, column=col_offset)

    # ── Save to bytes ──
    output = io.BytesIO()
    out_wb.save(output)
    output.seek(0)

    return {
        "output_bytes": output.getvalue(),
        "dates_processed": all_dates,
        "employee_count": len(employees),
        "record_count": record_count,
        "errors": errors,
    }


def build_dashboard_data(attendance_bytes: bytes, leave_bytes: Optional[bytes] = None) -> dict:
    """
    Parse raw attendance Excel (and optionally Leave/WFH Excel) and produce JSON data
    suitable for a dashboard (no persistence).

    Returns dict with:
      - 'dates_processed': list of date strings
      - 'employee_count': number of unique employees
      - 'record_count': total attendance records
      - 'errors': list of warning/error messages
      - 'employees': list of employee dashboards
    """
    wb = openpyxl.load_workbook(io.BytesIO(attendance_bytes), data_only=True)
    errors = []

    # Parse attendance sheets
    employees: dict[str, dict] = {}
    records: dict[str, dict[str, dict]] = {}
    all_dates: list[str] = []

    for sheet_name in wb.sheetnames:
        try:
            sheet_date = datetime.strptime(sheet_name.strip(), "%d-%m-%Y")
            date_str = sheet_date.strftime("%Y-%m-%d")
        except ValueError:
            errors.append(f"Skipped sheet '{sheet_name}': not a valid DD-MM-YYYY date")
            continue

        all_dates.append(date_str)
        ws = wb[sheet_name]

        for row in ws.iter_rows(min_row=8, values_only=True):
            if not row or len(row) < 8:
                continue

            emp_id_raw = row[1]
            if emp_id_raw is None:
                continue
            emp_id_str = str(emp_id_raw).strip()
            if not emp_id_str or emp_id_str.lower() in ("total", "grand total"):
                continue
            try:
                int(float(emp_id_str))
            except (ValueError, TypeError):
                continue

            emp_id = str(int(float(emp_id_str)))
            emp_name = str(row[2]).strip() if row[2] else ""
            if not emp_name:
                continue  # skip rows with no employee name
            department = str(row[3]).strip() if row[3] else "Unknown"

            in_time = parse_time_value(row[5])
            out_time = parse_time_value(row[6])
            total_time_str = parse_time_value(row[7])

            total_minutes = time_to_minutes(total_time_str)
            if total_minutes is None and in_time and out_time:
                in_min = time_to_minutes(in_time)
                out_min = time_to_minutes(out_time)
                if in_min is not None and out_min is not None and out_min > in_min:
                    total_minutes = out_min - in_min

            if emp_id not in employees:
                employees[emp_id] = {"name": emp_name, "department": department}

            if emp_id not in records:
                records[emp_id] = {}
            records[emp_id][date_str] = {
                "in_time": in_time or "",
                "out_time": out_time or "",
                "total_minutes": total_minutes,
            }

    wb.close()
    all_dates = sorted(set(all_dates))

    if not all_dates:
        return {
            "dates_processed": [],
            "employee_count": 0,
            "record_count": 0,
            "errors": errors or ["No valid date sheets found in the file"],
            "employees": [],
        }

    # Parse leave file (if provided)
    leave_records: dict[str, dict[str, str]] = {}
    leave_employees: dict[str, dict] = {}

    if leave_bytes:
        year = datetime.strptime(all_dates[0], "%Y-%m-%d").year
        leave_records, leave_employees, leave_errors = parse_leave_file(leave_bytes, year)
        errors.extend(leave_errors)

        # Merge employee info: prefer leave file names (cleaner), keep department from attendance
        for emp_id, leave_emp in leave_employees.items():
            if emp_id in employees:
                if leave_emp["name"] and leave_emp["name"] != "Unknown":
                    employees[emp_id]["name"] = leave_emp["name"]
            else:
                employees[emp_id] = {
                    "name": leave_emp["name"],
                    "department": leave_emp.get("department", "Unknown"),
                }

    weekend_flags = {d: is_weekend(d) for d in all_dates}

    # Build dashboard data
    record_count = 0
    employees_list = []

    sorted_emps = sorted(employees.keys(), key=lambda eid: int(eid))
    for emp_id in sorted_emps:
        emp = employees[emp_id]
        weekday_minutes = 0
        weekend_minutes = 0
        weekday_days = 0
        weekend_days = 0
        wfh_days = 0
        leave_days = 0
        absent_days = 0

        daily = []

        for date_str in all_dates:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            day_name = dt.strftime("%a")
            is_wkend = weekend_flags[date_str]
            rec = records.get(emp_id, {}).get(date_str)
            leave_status = leave_records.get(emp_id, {}).get(date_str)
            has_punch = rec and (rec["in_time"] or rec["out_time"])

            in_time = rec["in_time"] if rec else ""
            out_time = rec["out_time"] if rec else ""
            total_minutes = rec["total_minutes"] if rec else None
            total_hhmm = minutes_to_hhmm(total_minutes)

            status_label = ""
            status_type = ""
            note = ""
            leave_subtype = ""

            if leave_status == "WFH":
                record_count += 1
                wfh_days += 1
                weekday_minutes += 480
                total_minutes = 480
                total_hhmm = minutes_to_hhmm(480)
                status_label = "WFH"
                status_type = "wfh"
                in_time = ""
                out_time = ""

            elif leave_status in ("CL", "SL", "PL"):
                leave_days += 1
                status_label = leave_status
                status_type = "leave"
                leave_subtype = leave_status.lower()
                in_time = ""
                out_time = ""
                total_hhmm = ""
                total_minutes = None

            elif leave_status == "COMP OFF":
                leave_days += 1
                status_label = "COMP OFF"
                status_type = "comp_off"
                in_time = ""
                out_time = ""
                total_hhmm = ""
                total_minutes = None

            elif leave_status in ("1/2CL", "1/2SL", "1/2WFH", "1/2PL", "1/2COMP"):
                leave_days += 0.5
                status_label = leave_status
                status_type = "half_leave"
                half_subtype_map = {"1/2CL": "half_cl", "1/2SL": "half_sl", "1/2WFH": "half_wfh", "1/2PL": "half_pl", "1/2COMP": "half_comp"}
                leave_subtype = half_subtype_map.get(leave_status, "half_leave")
                if has_punch:
                    record_count += 1
                    mins = total_minutes if total_minutes and total_minutes > 0 else 0
                    weekday_minutes += mins
                    weekday_days += 1
                    total_hhmm = minutes_to_hhmm(mins)
                    note = "with_punch"
                else:
                    in_time = ""
                    out_time = ""
                    total_hhmm = ""
                    total_minutes = None

            elif leave_status == "HOLIDAY":
                status_label = "Holiday"
                status_type = "holiday"
                in_time = ""
                out_time = ""
                total_hhmm = ""
                total_minutes = None

            elif leave_status == "LWD":
                status_label = "LWD"
                status_type = "lwd"
                if has_punch:
                    record_count += 1
                    mins = total_minutes if total_minutes and total_minutes > 0 else 0
                    weekday_minutes += mins
                    weekday_days += 1
                    total_hhmm = minutes_to_hhmm(mins)
                else:
                    in_time = ""
                    out_time = ""
                    total_hhmm = ""
                    total_minutes = None

            elif has_punch:
                record_count += 1
                mins = total_minutes if total_minutes and total_minutes > 0 else 0
                total_hhmm = minutes_to_hhmm(mins)

                if is_wkend:
                    weekend_minutes += mins
                    weekend_days += 1
                    status_label = "Weekend Worked"
                    status_type = "weekend_worked"
                else:
                    weekday_minutes += mins
                    weekday_days += 1
                    status_label = "Present"
                    status_type = "present"

            elif is_wkend:
                status_label = "Holiday"
                status_type = "holiday"
                in_time = ""
                out_time = ""
                total_hhmm = ""
                total_minutes = None

            else:
                status_label = "Absent"
                status_type = "absent"
                absent_days += 1
                in_time = ""
                out_time = ""
                total_hhmm = ""
                total_minutes = None

            daily.append(
                {
                    "date": date_str,
                    "weekday": day_name,
                    "is_weekend": is_wkend,
                    "status": status_label,
                    "status_type": status_type,
                    "leave_subtype": leave_subtype,
                    "note": note,
                    "in_time": in_time,
                    "out_time": out_time,
                    "total_minutes": total_minutes,
                    "total_hhmm": total_hhmm,
                }
            )

        summary = {
            "weekday_minutes": weekday_minutes,
            "weekend_minutes": weekend_minutes,
            "total_minutes": weekday_minutes + weekend_minutes,
            "weekday_hours": minutes_to_hhmm(weekday_minutes),
            "weekend_hours": minutes_to_hhmm(weekend_minutes),
            "total_hours": minutes_to_hhmm(weekday_minutes + weekend_minutes),
            "working_days": weekday_days,
            "weekend_days": weekend_days,
            "wfh_days": wfh_days,
            "leave_days": leave_days,
            "absent_days": absent_days,
        }

        employees_list.append(
            {
                "emp_id": emp_id,
                "name": emp["name"],
                "department": emp["department"],
                "summary": summary,
                "daily": daily,
            }
        )

    return {
        "dates_processed": all_dates,
        "employee_count": len(employees),
        "record_count": record_count,
        "errors": errors,
        "employees": employees_list,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Google Sheets version of the dashboard builder
# Identical status/summary logic — only the data source changes.
# ─────────────────────────────────────────────────────────────────────────────

def _cell(row: list, idx: int):
    """Safe cell accessor — returns None if row is shorter than idx."""
    return row[idx] if idx < len(row) else None


def build_dashboard_data_from_google_sheets(
    att_sheet_id: str,
    leave_sheet_id: Optional[str],
    api_key: str,
) -> dict:
    """
    Fetch attendance + leave data from Google Sheets and produce the same
    dashboard JSON as build_dashboard_data().

    Args:
        att_sheet_id:   spreadsheet ID for the attendance sheet
        leave_sheet_id: spreadsheet ID for the leave/WFH sheet (optional)
        api_key:        Google Sheets API key

    Returns:
        Same structure as build_dashboard_data()
    """
    from backend.sheets_fetcher import get_sheet_tabs, get_sheet_values, batch_get_sheet_values

    errors: list[str] = []
    employees: dict[str, dict] = {}
    records: dict[str, dict[str, dict]] = {}
    all_dates: list[str] = []

    # ── Step 1: list all tabs in the attendance sheet ────────────────────────
    try:
        att_tabs = get_sheet_tabs(att_sheet_id, api_key)
    except Exception as e:
        return {
            "dates_processed": [],
            "employee_count": 0,
            "record_count": 0,
            "errors": [f"Cannot access attendance sheet: {e}"],
            "employees": [],
        }

    # ── Step 2: filter valid date tabs and batch-fetch all at once ───────────
    valid_tabs: list[tuple[str, str]] = []  # (tab_name, date_str)
    for tab in att_tabs:
        tab_name = tab["name"].strip()
        try:
            sheet_date = datetime.strptime(tab_name, "%d-%m-%Y")
            date_str = sheet_date.strftime("%Y-%m-%d")
            valid_tabs.append((tab_name, date_str))
        except ValueError:
            errors.append(f"Skipped tab '{tab_name}': not a valid DD-MM-YYYY date")

    # Batch fetch all attendance tabs (uses batchGet — ~10 API calls instead of ~365)
    att_tab_names = [t[0] for t in valid_tabs]
    all_att_data = batch_get_sheet_values(att_sheet_id, api_key, att_tab_names) if att_tab_names else {}

    # ── Step 3: parse each date tab from batch results ───────────────────────
    for tab_name, date_str in valid_tabs:
        all_dates.append(date_str)
        rows = all_att_data.get(tab_name, [])

        # Data starts at row 8 (0-indexed: index 7), same layout as Excel
        for row in rows[7:] if len(rows) > 7 else []:
            emp_id_raw = _cell(row, 1)
            if emp_id_raw is None or str(emp_id_raw).strip() == "":
                continue
            emp_id_str = str(emp_id_raw).strip()
            if emp_id_str.lower() in ("total", "grand total"):
                continue
            try:
                int(float(emp_id_str))
            except (ValueError, TypeError):
                continue

            emp_id = str(int(float(emp_id_str)))
            emp_name = str(_cell(row, 2) or "").strip()
            if not emp_name:
                continue  # skip rows with no employee name
            department = str(_cell(row, 3) or "Unknown").strip()

            in_time = parse_time_value(_cell(row, 5))
            out_time = parse_time_value(_cell(row, 6))
            total_time_str = parse_time_value(_cell(row, 7))

            total_minutes = time_to_minutes(total_time_str)
            if total_minutes is None and in_time and out_time:
                in_min = time_to_minutes(in_time)
                out_min = time_to_minutes(out_time)
                if in_min is not None and out_min is not None and out_min > in_min:
                    total_minutes = out_min - in_min

            if emp_id not in employees:
                employees[emp_id] = {"name": emp_name, "department": department}
            if emp_id not in records:
                records[emp_id] = {}
            records[emp_id][date_str] = {
                "in_time": in_time or "",
                "out_time": out_time or "",
                "total_minutes": total_minutes,
            }

    all_dates = sorted(set(all_dates))

    if not all_dates:
        return {
            "dates_processed": [],
            "employee_count": 0,
            "record_count": 0,
            "errors": errors or ["No valid date tabs (DD-MM-YYYY) found in attendance sheet"],
            "employees": [],
        }

    # ── Step 4: parse leave sheet ────────────────────────────────────────────
    leave_records: dict[str, dict[str, str]] = {}
    leave_employees: dict[str, dict] = {}
    year = datetime.strptime(all_dates[0], "%Y-%m-%d").year

    if leave_sheet_id:
        try:
            leave_tabs = get_sheet_tabs(leave_sheet_id, api_key)
        except Exception as e:
            errors.append(f"Cannot access leave sheet: {e}")
            leave_tabs = []

        # Filter valid month tabs and batch-fetch all at once
        valid_leave_tabs: list[tuple[str, int]] = []  # (tab_name, month_num)
        for tab in leave_tabs:
            clean_name = tab["name"].strip().lower()
            month_num = MONTH_MAP.get(clean_name)
            if month_num is None:
                errors.append(f"Leave sheet: skipped tab '{tab['name']}' (not a month name)")
            else:
                valid_leave_tabs.append((tab["name"], month_num))

        leave_tab_names = [t[0] for t in valid_leave_tabs]
        all_leave_data = batch_get_sheet_values(leave_sheet_id, api_key, leave_tab_names) if leave_tab_names else {}

        for tab_name, month_num in valid_leave_tabs:
            rows = all_leave_data.get(tab_name, [])
            if not rows:
                continue

            days_in_month = monthrange(year, month_num)[1]

            # Auto-detect header row (look for cell = 1, next cell = 2)
            header_row_idx = None
            day_col_start = None
            for row_idx, row in enumerate(rows[:5]):
                if not row:
                    continue
                for col_idx, val in enumerate(row):
                    if val is not None and str(val).strip():
                        try:
                            if int(float(str(val))) == 1:
                                nxt = _cell(row, col_idx + 1)
                                if nxt is not None:
                                    try:
                                        if int(float(str(nxt))) == 2:
                                            header_row_idx = row_idx
                                            day_col_start = col_idx
                                            break
                                    except (ValueError, TypeError):
                                        pass
                        except (ValueError, TypeError):
                            pass
                if header_row_idx is not None:
                    break

            if header_row_idx is None:
                header_row_idx = 0
                day_col_start = 3

            for row in rows[header_row_idx + 1:]:
                if not row or len(row) < day_col_start + 1:
                    continue

                emp_id_raw = _cell(row, 1)
                if emp_id_raw is None or str(emp_id_raw).strip() == "":
                    continue
                emp_id_str = str(emp_id_raw).strip()
                if emp_id_str.lower() in ("total", "grand total", "nan"):
                    continue
                try:
                    int(float(emp_id_str))
                except (ValueError, TypeError):
                    continue

                emp_id = str(int(float(emp_id_str)))
                emp_name = str(_cell(row, 2) or "").strip()
                if not emp_name:
                    continue  # skip rows with no employee name

                if emp_id not in leave_employees:
                    leave_employees[emp_id] = {"name": emp_name, "department": ""}
                if emp_id not in leave_records:
                    leave_records[emp_id] = {}

                for day in range(1, days_in_month + 1):
                    col_idx = day_col_start + (day - 1)
                    cell_val = _cell(row, col_idx)
                    status = _normalize_leave_status(cell_val)
                    if status in LEAVE_STATUSES:
                        leave_records[emp_id][f"{year}-{month_num:02d}-{day:02d}"] = status

        # Merge employee info
        for emp_id, leave_emp in leave_employees.items():
            if emp_id in employees:
                if leave_emp["name"] and leave_emp["name"] != "Unknown":
                    employees[emp_id]["name"] = leave_emp["name"]
            else:
                employees[emp_id] = {
                    "name": leave_emp["name"],
                    "department": leave_emp.get("department", "Unknown"),
                }

    # ── Step 4: build daily + summary (same logic as build_dashboard_data) ───
    weekend_flags = {d: is_weekend(d) for d in all_dates}
    record_count = 0
    employees_list = []

    sorted_emps = sorted(employees.keys(), key=lambda eid: int(eid))
    for emp_id in sorted_emps:
        emp = employees[emp_id]
        weekday_minutes = 0
        weekend_minutes = 0
        weekday_days = 0
        weekend_days = 0
        wfh_days = 0
        leave_days = 0
        absent_days = 0
        daily = []

        for date_str in all_dates:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            day_name = dt.strftime("%a")
            is_wkend = weekend_flags[date_str]
            rec = records.get(emp_id, {}).get(date_str)
            leave_status = leave_records.get(emp_id, {}).get(date_str)
            has_punch = rec and (rec["in_time"] or rec["out_time"])

            in_time = rec["in_time"] if rec else ""
            out_time = rec["out_time"] if rec else ""
            total_minutes = rec["total_minutes"] if rec else None
            total_hhmm = minutes_to_hhmm(total_minutes)

            status_label = ""
            status_type = ""
            note = ""
            leave_subtype = ""

            if leave_status == "WFH":
                record_count += 1
                wfh_days += 1
                weekday_minutes += 480
                total_minutes = 480
                total_hhmm = minutes_to_hhmm(480)
                status_label = "WFH"
                status_type = "wfh"
                in_time = ""
                out_time = ""

            elif leave_status in ("CL", "SL", "PL"):
                leave_days += 1
                status_label = leave_status
                status_type = "leave"
                leave_subtype = leave_status.lower()
                in_time = ""
                out_time = ""
                total_hhmm = ""
                total_minutes = None

            elif leave_status == "COMP OFF":
                leave_days += 1
                status_label = "COMP OFF"
                status_type = "comp_off"
                in_time = ""
                out_time = ""
                total_hhmm = ""
                total_minutes = None

            elif leave_status in ("1/2CL", "1/2SL", "1/2WFH", "1/2PL", "1/2COMP"):
                leave_days += 0.5
                status_label = leave_status
                status_type = "half_leave"
                half_subtype_map = {"1/2CL": "half_cl", "1/2SL": "half_sl", "1/2WFH": "half_wfh", "1/2PL": "half_pl", "1/2COMP": "half_comp"}
                leave_subtype = half_subtype_map.get(leave_status, "half_leave")
                if has_punch:
                    record_count += 1
                    mins = total_minutes if total_minutes and total_minutes > 0 else 0
                    weekday_minutes += mins
                    weekday_days += 1
                    total_hhmm = minutes_to_hhmm(mins)
                    note = "with_punch"
                else:
                    in_time = ""
                    out_time = ""
                    total_hhmm = ""
                    total_minutes = None

            elif leave_status == "HOLIDAY":
                status_label = "Holiday"
                status_type = "holiday"
                in_time = ""
                out_time = ""
                total_hhmm = ""
                total_minutes = None

            elif leave_status == "LWD":
                status_label = "LWD"
                status_type = "lwd"
                if has_punch:
                    record_count += 1
                    mins = total_minutes if total_minutes and total_minutes > 0 else 0
                    weekday_minutes += mins
                    weekday_days += 1
                    total_hhmm = minutes_to_hhmm(mins)
                else:
                    in_time = ""
                    out_time = ""
                    total_hhmm = ""
                    total_minutes = None

            elif has_punch:
                record_count += 1
                mins = total_minutes if total_minutes and total_minutes > 0 else 0
                total_hhmm = minutes_to_hhmm(mins)
                if is_wkend:
                    weekend_minutes += mins
                    weekend_days += 1
                    status_label = "Weekend Worked"
                    status_type = "weekend_worked"
                else:
                    weekday_minutes += mins
                    weekday_days += 1
                    status_label = "Present"
                    status_type = "present"

            elif is_wkend:
                status_label = "Holiday"
                status_type = "holiday"
                in_time = ""
                out_time = ""
                total_hhmm = ""
                total_minutes = None

            else:
                status_label = "Absent"
                status_type = "absent"
                absent_days += 1
                in_time = ""
                out_time = ""
                total_hhmm = ""
                total_minutes = None

            daily.append(
                {
                    "date": date_str,
                    "weekday": day_name,
                    "is_weekend": is_wkend,
                    "status": status_label,
                    "status_type": status_type,
                    "leave_subtype": leave_subtype,
                    "note": note,
                    "in_time": in_time,
                    "out_time": out_time,
                    "total_minutes": total_minutes,
                    "total_hhmm": total_hhmm,
                }
            )

        summary = {
            "weekday_minutes": weekday_minutes,
            "weekend_minutes": weekend_minutes,
            "total_minutes": weekday_minutes + weekend_minutes,
            "weekday_hours": minutes_to_hhmm(weekday_minutes),
            "weekend_hours": minutes_to_hhmm(weekend_minutes),
            "total_hours": minutes_to_hhmm(weekday_minutes + weekend_minutes),
            "working_days": weekday_days,
            "weekend_days": weekend_days,
            "wfh_days": wfh_days,
            "leave_days": leave_days,
            "absent_days": absent_days,
        }

        employees_list.append(
            {
                "emp_id": emp_id,
                "name": emp["name"],
                "department": emp["department"],
                "summary": summary,
                "daily": daily,
            }
        )

    return {
        "dates_processed": all_dates,
        "employee_count": len(employees),
        "record_count": record_count,
        "errors": errors,
        "employees": employees_list,
    }
